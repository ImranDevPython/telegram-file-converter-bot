import os
import tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def convert_csv_to_xlsx(csv_path: str) -> str:
    """
    Convert CSV file to XLSX format with formatting
    Args:
        csv_path (str): Path to the CSV file
    Returns:
        str: Path to the converted XLSX file
    """
    try:
        # Create temporary file for XLSX
        temp_xlsx = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_xlsx.close()

        # Read CSV file
        df = pd.read_csv(csv_path)
        
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        
        # Set sheet title (using CSV filename)
        title = os.path.splitext(os.path.basename(csv_path))[0]
        title = title.replace('_', ' ').replace('-', ' ').title()
        ws.title = title[:31]  # Excel sheet name length limit is 31 characters
        
        # Write headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            
            # Header style
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Data style
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
        
        # Adjust column widths
        for col in range(1, len(df.columns) + 1):
            column = get_column_letter(col)
            max_length = 0
            
            # Find maximum length in column
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Add some padding
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)  # Max width of 50
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        # Save the workbook
        wb.save(temp_xlsx.name)
        
        return temp_xlsx.name
        
    except Exception as e:
        if os.path.exists(temp_xlsx.name):
            os.remove(temp_xlsx.name)
        raise Exception(f"Error converting CSV to XLSX: {str(e)}") 